import os
import pandas as pd
from openpyxl import load_workbook


def xlsx_to_csv(xlsx):
    sheet = load_workbook(xlsx).worksheets[0]
    rows = list(sheet.rows)
    sheet_val = []
    for row in rows:
        row_val = [col.value for col in row]
        sheet_val.append(row_val)
    dt = pd.DataFrame(sheet_val[1:len(sheet_val)], columns=sheet_val[0])
    save_path = os.path.join(xlsx.replace("xlsx", "csv"))
    dt.to_csv(save_path, index=0)


def excel_to_csv_to_text(xlsx):
    sheet = load_workbook(xlsx).worksheets[0]
    rows = list(sheet.rows)
    sheet_val = []
    for row in rows:
        row_val = [col.value for col in row]
        sheet_val.append(row_val)
    # dt = pd.DataFrame(sheet_val[0:len(sheet_val) - 1], columns=sheet_val[0])
    dt = pd.DataFrame(sheet_val[1:len(sheet_val)], columns=sheet_val[0])
    save_path = os.path.join(xlsx.replace("xlsx", "csv"))
    dt.to_csv(save_path, index=0)
    text = ''
    with open(save_path) as f:
        for lin in f.readlines():
            text += lin.replace("_x000D_", "")
    return text


def unmerge_and_fill_cells(worksheet):
    all_merged_cell_ranges = list(
        worksheet.merged_cells.ranges
    )
    for merged_cell_range in all_merged_cell_ranges:
        merged_cell = merged_cell_range.start_cell
        worksheet.unmerge_cells(range_string=merged_cell_range.coord)
        for row_index, col_index in merged_cell_range.cells:
            cell = worksheet.cell(row=row_index, column=col_index)
            cell.value = merged_cell.value


def excel2txt(file_path):
    '''
    Convert the excel to text.
    :param file_path:
    :return:str
    '''
    all_table_text = ''
    wb = load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        unmerge_and_fill_cells(sheet)

    head_info = []  # TODU
    for index, each in enumerate(sheet.iter_cols()):
        value = each[0].value
        if value:
            value = each[0].value.replace('_x000D_', '').strip()
        else:
            value = 'unavailable'
        head_info.append(value)
    maxrows = sheet.max_row
    for i in range(1, maxrows):  # table head maybe row=1
        text = []
        for index, each in enumerate(sheet.iter_cols()):
            value = each[i].value
            if value:
                value = each[i].value.replace('_x000D_', '')
            else:
                value = "unavailable"
            text.append(value)
        merge_text = []
        for j in range(len(head_info)):
            merge_text.append(head_info[j])
            merge_text.append(text[j])
        merge_text = " ".join(merge_text)
        all_table_text = all_table_text + merge_text + "\n"
    return all_table_text


if __name__ == '__main__':
    excel = '/Users/57block/PycharmProjects/OPGEE/opgee/demos/output/ExtractTextTableWithFigureTableRendition/tables/fileoutpart19.xlsx'

    print(excel_to_csv_to_text(excel))
