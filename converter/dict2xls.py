import os
import sys
import pandas as pd
from pathlib import Path
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill
import numpy as np
from typing import Optional, Union, Any, List
from collections import OrderedDict

from config import SPLIT_REFERENCE_CHAR, REPLACE_SIGN, SPLIT_BLOCK_CHAR
from eval.xls_parser import ParseExcel
from eval.parser.parser_result import SUCCESS, ERROR
from schema.variables import get_units, is_gt_variable
from lib.my_logger import logger
from lib.helper import is_numeric


# Anjing: This is not the best practice, but
# (1) we would like run every file individually
# without going back and forth with python -m
# (2) we do not want to maintain PYTHON_PATH
# especially when we work on mutliple projects
# at the same time
# (3) we lost the benefit of editor auto completion
# but that's okay
this_file_path = os.path.dirname(os.path.realpath(__file__))
sys.path.insert(0, this_file_path + '/../')


def write_df_to_excel(filepath: str, dataframe: pd.DataFrame, sheet: str = 'Sheet1', create: bool = False, if_sheet_exists='error'):
    if not (parent_dir := Path(filepath).parent).exists():
        os.makedirs(str(parent_dir))

    if create:
        mode = 'w'
        if_sheet_exists = None
    else:
        mode = 'a'

    with pd.ExcelWriter(filepath, engine='openpyxl', mode=mode, if_sheet_exists=if_sheet_exists) as writer:
        dataframe.to_excel(writer, sheet_name=sheet)


def convert_dict_to_xls(result_folder_path: str,
                        dt: OrderedDict,
                        dt_errors=None,
                        ground_truth: Optional[ParseExcel] = None,
                        success_only: bool = False,
                        ):
    """
       Converts ordered dictionary to simple Excel spreadsheet
       :param dt_errors:
       :param dt: data ordered dictionary
       :param result_folder_path: Excel spreadsheet output filepath
       :param ground_truth: Ground Truth excel object for extracted paper
       :return: None
       """
    value_file_path = Path(result_folder_path, 'extract.xlsx')
    ref_file_path = Path(result_folder_path, 'extract_ref.xlsx')
    raw_file_path = Path(result_folder_path, 'extract_raw.xlsx')

    df_value = convert_dict_to_dataframe(dt, mode='value',
                                   success_only=success_only)
    df_reference = convert_dict_to_dataframe(dt, mode='ref',
                                    success_only=success_only)
    df_raw = convert_dict_to_dataframe(dt, mode='raw')

    if dt_errors:
        df_errors = convert_dict_to_dataframe(dt_errors, mode='raw')
        xls_fill_color(value_file_path.as_posix(), df_errors)
        xls_fill_color(ref_file_path.as_posix(), df_errors)
    else:
        write_df_to_excel(value_file_path.as_posix(), df_value, create=True)
        write_df_to_excel(ref_file_path.as_posix(), df_reference, create=True)
        write_df_to_excel(raw_file_path.as_posix(), df_raw, create=True)
        if ground_truth is not None:
            ground_truth.to_excel(value_file_path.as_posix(), sheet='Ground Truth')
            ground_truth.to_excel(ref_file_path.as_posix(), sheet='Ground Truth')
            ground_truth.to_excel(raw_file_path.as_posix(), sheet='Ground Truth')


def fill_color_to_extract_and_ref(result_folder_path: str,
                                  dt_errors: OrderedDict):
    ''' fill color to extract and extract_ref sheet '''
    value_file_path = Path(result_folder_path, 'extract.xlsx')
    ref_file_path = Path(result_folder_path, 'extract_ref.xlsx')

    df_errors = convert_dict_to_dataframe(dt_errors)
    xls_fill_color(value_file_path, df_errors)
    xls_fill_color(ref_file_path, df_errors)


def convert_dict_to_dataframe(dt: Union[OrderedDict, dict],
                              mode: str = None,
                              success_only: bool = False):
    tmp_data = OrderedDict()
    fields = list(dt.keys())
    index = list(dt[fields[0]].keys())
    for field in fields:
        tmp_data[field] = []
        for var, values in dt[field].items():
            if mode == 'value':
                # If success_only is set,
                # values with error parse status would be set to NaN.
                if check_to_replace_value(success_only, values):
                    value = np.nan
                else:
                    value = values[0]

                if is_gt_variable(var):
                    tmp_data[field].append(value)
                else:
                    if var in index:
                        index.remove(var)

            elif mode == 'ref':
                # If success_only is set, ref with error parse status
                # would add a '#' sign in the front.
                if check_to_replace_value(success_only, values):
                    ref = ' '.join([REPLACE_SIGN, _get_ref(values)])
                else:
                    ref = _get_ref(values)

                if is_gt_variable(var):
                    tmp_data[field].append(ref)
                else:
                    if var in index:
                        index.remove(var)

            elif mode == 'raw':
                tmp_data[field].append(values[3])
            else:
                tmp_data[field].append(values)
    tmp_data['unit'] = get_units(index)
    tmp_data.move_to_end('unit', last=False)

    df = pd.DataFrame(tmp_data, index=index, columns=list(tmp_data.keys()))
    return df


def _get_ref(values: List[Any]) -> str:
    if isinstance(values[0], float | int) and np.isnan(values[0]):
        return values[2]
    else:
        if values[2] != '':
            return f"{str(values[0])} {str(values[1])} {SPLIT_REFERENCE_CHAR}{values[2]}"
        else:
            return f"{str(values[0])} {str(values[1])}"


def xls_fill_color(filepath, df_errors):
    '''
    False positive (should be negative) red / error value 1
    False positive (wrong value) orange / error value 3
    False negative (missing value) blue / error value 2
    '''
    if not Path(filepath).exists():
        err_msg = f"file {filepath} does not exists, no color filled"
        logger.error(err_msg)
        return

    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFFCCCB',
                           fill_type='solid', patternType='darkUp')
    blue_fill = PatternFill(start_color='FF46BDC6', end_color='FF91D2FF',
                            fill_type='solid', patternType='darkUp')
    orange_fill = PatternFill(start_color='FFFFA500', end_color='FFFFA500',
                              fill_type='solid', patternType='darkUp')
    wb = load_workbook(filepath)

    ws = wb.get_sheet_by_name('Sheet1')
    for row_ind, (index, row_data) in enumerate(df_errors.iterrows(), start=2):
        for col_ind, error_value in enumerate(row_data, start=2):
            if error_value == 1:
                ws.cell(row=row_ind, column=col_ind).fill = red_fill
            if error_value == 2:
                ws.cell(row=row_ind, column=col_ind).fill = blue_fill
            if error_value == 3:
                ws.cell(row=row_ind, column=col_ind).fill = orange_fill
    wb.save(filepath)


def check_to_replace_value(success_only: bool, values: List)-> bool:
    '''
    decide whether we need to give up the parsed value
    and replace it with a NaN
    '''

    # if extracted is a pure value, then we don't need to replace
    # and try our luck
    # there are many cases LLM just doesn't return a unit along with the value
    if is_numeric(str(values[0])) is True:
        return False

    # if extracted value is an error, then we give up and replace it
    return success_only and \
            (values[4] == ERROR or SPLIT_BLOCK_CHAR in str(values[0]))
