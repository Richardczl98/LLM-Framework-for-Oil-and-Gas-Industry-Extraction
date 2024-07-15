import os
import pandas as pd
from pathlib import Path
from typing import List, Optional, Set
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from collections import OrderedDict
import sys
import json


this_file_path = os.path.dirname(os.path.realpath(__file__))
sys.path.insert(0, this_file_path + '/../')

from config import ROOT_DIR, RESULT_DIR, RESULT_DIFF_DIR, AGGREGATE_COLUMN_NAME_SPLITTER
from converter.dict2xls import write_df_to_excel
from lib.file_op import get_last_dir
from lib.my_logger import logger

INDEX_COLS = ['Unnamed: 0', 'Unit']


def merge_results(file_name: str, result_dir: Optional[str | Path] = RESULT_DIR, create=False):
    if 'ref' in file_name:
        sheet = 'ref'
    elif 'raw' in file_name:
        sheet = 'raw'
    else:
        sheet = 'Sheet1'

    summary_path = Path(result_dir) / 'extract_aggr.xlsx'
    results_last_dirs = get_last_dir(*os.path.split(result_dir))
    merged_df = pd.DataFrame()
    merged_json = {}
    color_dict = dict()
    for extracted_dir, paper in results_last_dirs:
        excel_file = Path(extracted_dir, file_name)
        extracted_df = pd.read_excel(str(excel_file), sheet_name='Sheet1')
        extracted_df.rename(columns={'unit': 'Unit'}, inplace=True)
        extracted_df.set_index(INDEX_COLS, inplace=True)
        extracted_df.rename(
            columns={old: AGGREGATE_COLUMN_NAME_SPLITTER.join([paper, old]) for old in extracted_df.columns},
            inplace=True
        )
        ground_truth_df = pd.read_excel(str(excel_file), sheet_name='Ground Truth')
        ground_truth_df.set_index(INDEX_COLS, inplace=True)
        ground_truth_df.rename(
            columns={old: AGGREGATE_COLUMN_NAME_SPLITTER.join([paper, old, 'GT']) for old in ground_truth_df.columns},
            inplace=True
        )
        load_color_in_excel(os.path.join(extracted_dir, 'extract_ref.xlsx'), color_dict, extracted_df.columns)
        for gt_col, pred_col in zip(ground_truth_df.columns, extracted_df.columns):
            merged_df = pd.concat([merged_df, ground_truth_df[gt_col], extracted_df[pred_col]], axis=1)
        merged_df.index.set_names(INDEX_COLS, inplace=True)

        with open(os.path.join(extracted_dir, 'extract.json'), 'r') as fp:
            merged_json[paper] = json.load(fp)
    write_df_to_excel(filepath=str(summary_path), dataframe=merged_df, sheet=sheet, create=create)
    if sheet == 'ref':
        with open(str(summary_path).replace('xlsx', 'json'), 'w') as fp:
            json.dump(merged_json, fp)
        logger.info(f'Successfully save merged {sheet} table in {str(summary_path).replace("xlsx", "json")}')

    fill_merged_color(summary_path, color_dict, list(merged_df.columns), sheet=sheet)
    logger.info(f'Successfully save merged {sheet} table in {summary_path}')


def load_color_in_excel(extract_dir: Optional[str | Path], color_dict: dict, columns: List[str], sheet: str = 'Sheet1'):
    wb = load_workbook(extract_dir)
    ws = wb[sheet]
    variables = ws['A']

    for col, col_name in zip(list(ws.iter_cols())[2:], columns):
        col_colors = OrderedDict()
        for var, cell in zip(variables, col):
            color = PatternFill(
                start_color=cell.fill.start_color.index,
                end_color=cell.fill.end_color.index,
                fill_type=cell.fill.fill_type,
            )
            col_colors[var.value] = color
        color_dict.update({col_name: col_colors})


def fill_merged_color(file_path: Optional[str | Path], color_dict: dict, columns: List[str], sheet: str):
    wb = load_workbook(file_path)
    ws = wb[sheet]
    for col_name, rows in color_dict.items():
        column_letter = get_column_letter(columns.index(col_name) + 3)      # add variable column and unit column
        for row_id, pattern in enumerate(rows.values()):
            coordinate = column_letter + str(row_id + 1)
            cell = ws[coordinate]
            cell.fill = pattern

    wb.save(file_path)


def compare_multi_results(rslt_paths: List[str | Path]):
    """
    Compare ref table for different result batches.
    This comparison would based on the 'extract_aggr.xlsx' table for each result batch.
    Before run this comparison please generate 'extract_aggr.xlsx' file for each result batch first.
    This comparison would only consider fields included in all batches.
    The compared result would be saved in 'result_diff/extract_diff.xlsx' file.
    Only variables with different value be save in the result file.

    :param rslt_paths: The path of 'extract_aggr.xlsx' for each result batch.
    """

    diff_path = RESULT_DIFF_DIR / 'extract_diff.xlsx'
    aggr_df_list = [pd.read_excel(Path(ROOT_DIR, path, 'extract_aggr.xlsx'), sheet_name='ref').set_index(INDEX_COLS) for path in rslt_paths]
    aggr_color_list = [{} for _ in rslt_paths]
    for path, aggr_color, columns in zip(rslt_paths, aggr_color_list, aggr_df_list):
        load_color_in_excel(Path(ROOT_DIR, path, 'extract_aggr.xlsx'), aggr_color, columns, sheet='ref')

    # Here we only consider fields exist in all batch of results.
    common_columns = _get_common_columns(aggr_df_list)
    gt_columns = _get_gt_columns(common_columns)
    field_columns = _get_field_columns(gt_columns)
    papers = _get_papers(gt_columns)

    compare_df_by_paper = {}
    for paper in papers:
        compare_df = pd.DataFrame()
        compare_color = {}
        fields_cols_in_paper = _get_fields_by_paper(paper, field_columns)
        gt_in_paper = _get_fields_by_paper(paper, gt_columns)
        field_names = _get_field_name_from_columns(fields_cols_in_paper)

        for old_gt, field, old_field_col in zip(gt_in_paper, field_names, fields_cols_in_paper):
            new_gt_name = _generate_new_gt_col_name(field)
            for path, aggr_df, aggr_color in zip(rslt_paths, aggr_df_list, aggr_color_list):
                new_field_name = _generate_new_field_col_name(path, field)

                compare_df[new_gt_name] = aggr_df[old_gt]
                compare_df[new_field_name] = aggr_df[old_field_col]
                compare_color[new_field_name] = aggr_color.get(old_field_col)

        _compare(compare_df, compare_color)
        compare_df_by_paper.update({paper: compare_df})

        create = True if not diff_path.exists() else False
        write_df_to_excel(filepath=str(diff_path), dataframe=compare_df, sheet=paper,
                          create=create, if_sheet_exists='replace')
        fill_merged_color(diff_path, compare_color, list(compare_df.columns), sheet=paper)
        logger.info(f'eval_diff.xlsx has been generated in result_diff dir')


def _generate_new_field_col_name(batch_name: str, field_name: str) -> str:
    return AGGREGATE_COLUMN_NAME_SPLITTER.join([batch_name, field_name])


def _generate_new_gt_col_name(field_name: str) -> str:
    return AGGREGATE_COLUMN_NAME_SPLITTER.join([field_name, 'GT'])


def _get_unique_columns(aggr_df_list: List[pd.DataFrame]):
    columns = []
    for aggr_df in aggr_df_list:
        [columns.append(col) for col in aggr_df.columns if col not in columns]

    return columns


def _get_common_columns(aggr_df_list: List[pd.DataFrame]):
    all_cols = list(aggr_df_list[0].columns)
    common_columns = set(aggr_df_list[0].columns)
    for aggr_df in aggr_df_list[1:]:
        common_columns = common_columns.intersection(aggr_df.columns)
        all_cols += list(aggr_df.columns)

    return sorted(common_columns, key=lambda x: all_cols.index(x))


def _get_gt_columns(columns: Optional[List | Set]):
    return [col for col in columns if 'GT' in col]


def _get_field_columns(columns: Optional[List | Set]):
    return [col.rsplit(AGGREGATE_COLUMN_NAME_SPLITTER, 1)[0] for col in columns if 'GT' in col]


def _get_field_name_from_columns(columns: Optional[List | Set]):
    return [col.rsplit(AGGREGATE_COLUMN_NAME_SPLITTER)[1] for col in columns]


def _get_papers(columns: Optional[List | Set]):
    papers = []
    for col in columns:
        if (paper := col.split(AGGREGATE_COLUMN_NAME_SPLITTER)[0]) not in papers:
            papers.append(paper)

    return papers


def _get_fields_by_paper(paper: str, field_cols: Optional[List | Set]):
    return [field for field in field_cols if paper in field]


def _compare(compare_df: pd.DataFrame, compare_color: dict):
    equal_vars = []
    for index in compare_df.index:
        if _compare_equal(compare_df.loc[index]):
            equal_vars.append(index)
    compare_df.drop(labels=equal_vars, axis=0, inplace=True)

    for field_color in compare_color.values():
        for var in equal_vars:
            field_color.pop(var[0])


def _compare_equal(row: pd.Series):
    aggr_fields_col = [col for col in row.index if 'GT' not in col]

    aggr_field_equal = []
    for i in range(0, len(aggr_fields_col), 2):
        aggr_field_values = []
        for col in aggr_fields_col[i:i+2]:
            if pd.notna(row[col]):
                aggr_field_values.append(row[col].split('@')[0].strip())
            else:
                aggr_field_values.append(row[col])

        if len(set(aggr_field_values)) == 1:
            aggr_field_equal.append(True)
        else:
            aggr_field_equal.append(False)

    return all(aggr_field_equal)


if __name__ == '__main__':
    pass
