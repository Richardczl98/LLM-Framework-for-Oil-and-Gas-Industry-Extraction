# -*- coding: utf-8 -*-
# Author: Anjing Wang
# Date: Oct-19-2023

"""file operation related helper functions"""

import os
import sys
import csv
import json

this_file_path = os.path.dirname(os.path.realpath(__file__))
sys.path.insert(0, this_file_path + '/../')

from lib.my_logger import logger
from lib.helper import get_time_str
import openpyxl
from openpyxl.styles import PatternFill
from pathlib import Path
import pandas as pd
from collections import Counter
from traceback import format_exc


def save_dict_to_json(data, filename):
    """
    Save a dictionary to a JSON file.

    :param data: Dictionary to save.
    :param filename: File name for the JSON output.
    """
    try:
        with open(filename, 'w', encoding="utf-8") as file:
            json.dump(data, file, indent=4)
        print(f"Data successfully saved to {filename}")
    except IOError as e:
        print(f"Error saving to file: {e}")


def read_json_to_dict(filename):
    """
    Read a JSON file and convert it to a dictionary.

    :param filename: File name of the JSON to read.
    :return: A dictionary containing the data from the JSON file.
    """
    try:
        with open(filename, 'r', encoding="utf-8") as file:
            data = json.load(file)
            print(f"Data successfully read from {filename}")
            return data
    except FileNotFoundError:
        print(f"No such file named {filename}")
        return {}
    except json.JSONDecodeError:
        print(f"File {filename} is not a valid JSON.")
        return {}
    except IOError as e:
        print(f"Error reading file: {e}")
        return {}


def write_to_file(content: str, file_path: str) -> bool:
    """ Always overwirte the file """
    try:
        with open(file_path, 'w', encoding="utf-8") as f:
            f.write(content)
        return True
    except Exception as e:
        print(f"Error writing to file: {e}")
        return False


def write_to_csv(csv_file: str, header: list, data: list) -> bool:
    try:
        with open(csv_file, mode='w', newline='') as file:
            writer = csv.writer(file)
            if header:
                writer.writerow(header)
            for row in data:
                writer.writerow(row)
        return True
    except Exception as e:
        logger.error(f"Error writing to file: {e}")
        return False


def read_csv(path: str) -> list:
    with open(path, mode='r', newline='') as file:
        reader = csv.reader(file)
        data = list(reader)
    return data


def get_last_dir(base_dir, result_relative_dir):
    split_char = os.sep
    full_path = os.path.join(base_dir, result_relative_dir)
    level = len(result_relative_dir.split(split_char))

    if level > 2:
        raise Exception("Too many folder paths, only level 2 is supported")

    if level == 1:
        result = []
        sub_dirs = {os.path.join(full_path, name): [sec for sec in os.listdir(os.path.join(full_path, name)) if
                                                    os.path.isdir(os.path.join(full_path, name, sec))] for name in
                    os.listdir(full_path) if
                    os.path.isdir(os.path.join(full_path, name))}

        for k, v in sub_dirs.items():
            result.append([os.path.join(k, max(v)), k.split(split_char)[-1:][0]])

        return result

    else:
        sub_dirs = [name for name in os.listdir(full_path) if os.path.isdir(os.path.join(full_path, name))]
        if sub_dirs:
            return [[os.path.join(full_path, max(sub_dirs)), full_path.split(split_char)[-1:][0]]]
        return []


def generate_percent_csv(path: str, data: list):
    data = data[1:]
    sums = sum(data)
    result = [['TP', 'FP']]
    result.append([data[0], data[1]])
    result.append([f'{round(data[0] / sums * 100, 2)}%', f'{round(data[1] / sums * 100, 2)}%'])
    result.append(['FN', 'TN'])
    result.append([data[3], data[2]])
    result.append([f'{round(data[3] / sums * 100, 2)}%', f'{round(data[2] / sums * 100, 2)}%'])
    write_to_csv(path, [], result)


def generate_percent_without_tn_csv(path: str, data: list):
    data = data[1:]
    sums = sum(data) - float(data[2])
    result = [['TP', 'FP']]
    result.append([data[0], data[1]])
    result.append([f'{round(data[0] / sums * 100, 2)}%', f'{round(data[1] / sums * 100, 2)}%'])
    result.append(['FN'])
    result.append([data[3]])
    result.append([f'{round(data[3] / sums * 100, 2)}%'])
    write_to_csv(path, [], result)


def add_cost_data(path: str, cost_list: list):
    data = read_json_to_dict(path)
    cost_list.append([data.get('cost', ""), data.get('time', "")])


def convert_time(data):
    data[2] = round(data[2])
    hour = data[2] // 3600
    min = data[2] % 3600 // 60
    sec = data[2] % 3600 % 60
    data[2] = f'{hour}hr{min}min{sec}sec'
    pass


def summary_result(data: list):
    return [sum(float(num) for num in nums if num) for nums in zip(*data)]


def combine_data(item1: list, item2: list):
    return [[item1[idx]] + item2[idx] for idx in range(len(item2))]


def sum_dataframe(df: pd.DataFrame, summed_df: pd.DataFrame = None) -> pd.DataFrame:
    '''
     Merge results for each article
    :param df: DataFrame
    :param summed_df: DataFrame
    :return: DataFrame
    '''
    summed_df = pd.DataFrame(index=df.index) if summed_df.empty else summed_df

    for col in df.columns:
        if col not in summed_df.columns:
            summed_df[col] = pd.Series(dtype='object')

    for index, row in df.iterrows():
        for col, value in row.items():
            if pd.notnull(value):
                if pd.isnull(summed_df.loc[index, col]):
                    summed_df.loc[index, col] = value
                else:
                    try:
                        value_dict = eval(value)
                        if pd.isna(summed_df.at[index, col]):
                            summed_df.at[index, col] = value_dict
                        else:
                            org_dict = eval(summed_df.at[index, col])
                            counter = Counter(org_dict) + Counter(value_dict)
                            summed_df.at[index, col] = str(dict(counter))
                    except:
                        try:
                            value = float(value)
                            if pd.isna(summed_df.at[index, col]):
                                summed_df.loc[index, col] = value
                            else:
                                summed_df.at[index, col] += value
                        except:
                            # value is a string
                            summed_df.loc[index, col] = value

    return summed_df


def summed_result(df: pd.DataFrame, summed_dic: dict) -> pd.DataFrame:
    '''
    Count the value of each row index
    df: the dataframe
    summed_dic: the dictionary
    return: the dataframe
    '''
    for index, row in df.iterrows():
        for col, value in row.items():
            try:
                dic = eval(value)
                tmp = summed_dic.get(index, {})
                tmp = Counter(tmp) + Counter(dic)
                summed_dic[index] = tmp
            except:
                pass

    for key, colors_counter in summed_dic.items():
        if key in df.index:
            for color, count in colors_counter.items():
                df.loc[key, color] = count

    return df


def generate_summary_csv(path: str, start_date: str = None) -> list:
    '''
    path: the path of the result like: result
    start_date: the start time format '%y%m%d' 231019
    '''
    if os.path.isabs(path):
        base_path, path = os.path.split(path)
    else:
        base_path = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))

    full_path = os.path.join(base_path, path)

    data_list = []
    header = []
    folder_list = []
    cost_folder_list = []
    cost_list = []
    time_str = get_time_str()
    old_data_list = []
    old_folder_list = []
    summed_df = pd.DataFrame()
    summed_dic = {}

    SPLIT_CHAR = os.sep
    level = len(path.split(SPLIT_CHAR))
    if level > 2:
        raise Exception("Too many folder paths, only level 2 is supported")

    # TP,FP,TN,FN
    try:
        if not start_date:
            dirs = get_last_dir(base_path, path)
            for dir in dirs:
                for file in os.listdir(dir[0]):
                    if file == 'eval_matrix.csv':
                        folder_list.append(dir[1])
                        data = read_csv(os.path.join(dir[0], file))
                        header.extend(data[0]) if not header else header
                        data_list.extend(data[1:])

                    if file == "stats.json":
                        cost_folder_list.append(dir[1])
                        add_cost_data(os.path.join(dir[0], file), cost_list)

                    if file == 'eval_old_matrix.csv':
                        old_folder_list.append(dir[1])
                        data = read_csv(os.path.join(dir[0], file))
                        old_data_list.extend(data[1:])

                    if file == "eval_field_errors.xlsx":
                        df = pd.read_excel(os.path.join(dir[0], file), index_col=0)
                        summed_df = sum_dataframe(df, summed_df)

        else:
            if level == 1:
                for dir in os.listdir(full_path):
                    if os.path.isdir(os.path.join(full_path, dir)):
                        for second_dir in os.listdir(os.path.join(full_path, dir)):
                            # result_time >= start
                            if second_dir[:6] >= start_date:
                                if os.path.isdir(os.path.join(full_path, dir, second_dir)):
                                    for file in os.listdir(os.path.join(full_path, dir, second_dir)):
                                        if file == 'eval_matrix.csv':
                                            folder_list.append(dir)
                                            data = read_csv(os.path.join(full_path, dir, second_dir, file))
                                            header.extend(data[0]) if not header else header
                                            data_list.extend(data[1:])

                                        if file == "stats.json":
                                            cost_folder_list.append(dir)
                                            add_cost_data(os.path.join(full_path, dir, second_dir, file), cost_list)

                                        if file == 'eval_old_matrix.csv':
                                            old_folder_list.append(dir)
                                            data = read_csv(os.path.join(full_path, dir, second_dir, file))
                                            old_data_list.extend(data[1:])

                                        if file == "eval_field_errors.xlsx":
                                            df = pd.read_excel(os.path.join(full_path, dir, second_dir, file), index_col=0)
                                            summed_df = sum_dataframe(df, summed_df)

            else:
                for second_dir in os.listdir(os.path.join(full_path)):
                    if second_dir[:6] >= start_date:
                        if os.path.isdir(os.path.join(full_path, second_dir)):
                            for file in os.listdir(os.path.join(full_path, second_dir)):
                                if file == 'eval_matrix.csv':
                                    folder_list.append(full_path.split(SPLIT_CHAR)[-1:][0])
                                    data = read_csv(os.path.join(full_path, second_dir, file))
                                    header.extend(data[0]) if not header else header
                                    data_list.extend(data[1:])

                                if file == "stats.json":
                                    cost_folder_list.append(full_path.split(SPLIT_CHAR)[-1:][0])
                                    add_cost_data(os.path.join(full_path, second_dir, file), cost_list)

                                if file == 'eval_old_matrix.csv':
                                    old_folder_list.append(full_path.split(SPLIT_CHAR)[-1:][0])
                                    data = read_csv(os.path.join(full_path, second_dir, file))
                                    old_data_list.extend(data[1:])

                                if file == "eval_field_errors.xlsx":
                                    df = pd.read_excel(os.path.join(full_path, second_dir, file), index_col=0)
                                    summed_df = sum_dataframe(df, summed_df)

        if data_list:
            result = summary_result(data_list)
            data_list = combine_data(folder_list, data_list)
            data_list.extend([""])
            header.insert(0, "")
            result.insert(0, "Total")
            data_list.append(result)
            generate_percent_csv(os.path.join(full_path, f'{time_str}_percent.csv'), result)
            generate_percent_without_tn_csv(os.path.join(full_path, f'{time_str}_percent_without_tn.csv'), result)
            write_to_csv(os.path.join(full_path, f'{time_str}_eval_matrix.csv'), header, data_list)

        if cost_list:
            cost_result = summary_result(cost_list)
            cost_list = combine_data(cost_folder_list, cost_list)
            cost_result.insert(0, "Total")
            cost_header = ["", "Cost", "Time"]
            cost_list.append(cost_result)
            [convert_time(row) for row in cost_list]
            write_to_csv(os.path.join(full_path, f'{time_str}_eval_cost.csv'), cost_header, cost_list)
            pass

        if not summed_df.empty:
            summed_df = summed_result(summed_df, summed_dic)
            summed_df.to_excel(os.path.join(full_path, f'{time_str}_eval_field_errors.xlsx'), index=True)

    except Exception as e:
        logger.error(f"Error generate csv from file: str({e})")
        logger.error(f"{format_exc()}")
        return False

    logger.info(f"Successfully generate statistics csv")
    return True


def combined_excel(excel_files: list):
    new_wb = openpyxl.Workbook()
    new_wb.remove(new_wb.active)
    time_str = get_time_str()

    BASE_DIRECTORY = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))

    try:
        for file_index, excel_path in enumerate(excel_files):
            excel_file = os.path.join(BASE_DIRECTORY, excel_path)
            if not os.path.exists(excel_file):
                raise Exception(f"{excel_file} doesn't exist")

            parts = Path(excel_file).parts
            sheet_title = parts[-2]
            excel_name = parts[-3]
            store_path = Path(*parts[:-3])

            src_wb = openpyxl.load_workbook(excel_file, data_only=True)

            src_ws = src_wb['Sheet1']

            new_ws = new_wb.create_sheet(title=sheet_title)

            for row in src_ws.iter_rows():
                for cell in row:
                    new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)

                    if cell.fill.start_color.index != '00000000':
                        fill_color = PatternFill(start_color=cell.fill.start_color.index,
                                                 end_color=cell.fill.start_color.index, fill_type='solid')
                        new_cell.fill = fill_color
    except Exception as e:
        logger.error(f'Error generate combined excel:{e}')
        return

    #
    store_excel_path = os.path.join(store_path, f'{time_str}-combined-{excel_name}.xlsx')
    new_wb.save(store_excel_path)
    logger.info(f"Successfully combined excel '{store_excel_path}'")


if __name__ == '__main__':
    pass